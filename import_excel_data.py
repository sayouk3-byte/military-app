import os
import sys
import sqlite3
import openpyxl
from datetime import datetime, date

# Ensure UTF-8 console output
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE_DIR, 'military_db.sqlite')
EXCEL_FILE = os.path.join(BASE_DIR, 'Personnel_Data.xlsx')
PHOTOS_DIR = os.path.join(BASE_DIR, 'photos')

def format_date_str(val):
    if val is None:
        return ''
    if isinstance(val, (datetime, date)):
        return val.strftime('%d/%m/%Y')
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ('none', 'null', '-'):
        return ''
    return val_str

def parse_pob(pob_str):
    if not pob_str:
        return '', '', '', ''
    parts = [p.strip() for p in str(pob_str).split('-') if p.strip()]
    if len(parts) == 4:
        return parts[0], parts[1], parts[2], parts[3]
    elif len(parts) == 3:
        return '', parts[0], parts[1], parts[2]
    elif len(parts) == 2:
        return '', '', parts[0], parts[1]
    elif len(parts) == 1:
        return '', '', '', parts[0]
    else:
        return parts[0], parts[1], parts[2], parts[-1]

def parse_addr(addr_str):
    if not addr_str:
        return '', '', '', '', '', ''
    parts = [p.strip() for p in str(addr_str).split('-') if p.strip()]
    house, group, village, commune, district, province = '', '', '', '', '', ''
    if len(parts) == 6:
        house, group, village, commune, district, province = parts
    elif len(parts) == 5:
        if 'ផ្ទះ' in parts[0]:
            house, village, commune, district, province = parts
        elif 'ក្រុម' in parts[0]:
            group, village, commune, district, province = parts
        else:
            village, commune, district, province = parts[1], parts[2], parts[3], parts[4]
    elif len(parts) == 4:
        village, commune, district, province = parts
    elif len(parts) == 3:
        commune, district, province = parts
    elif len(parts) == 2:
        district, province = parts
    elif len(parts) == 1:
        province = parts[0]
    else:
        province = parts[-1]
    return house, group, village, commune, district, province

def find_photo_path(id_card, photo_type='personal'):
    if not id_card:
        return ''
    id_clean = str(id_card).strip()
    id_padded = id_clean.zfill(6)
    
    candidates = []
    if photo_type == 'personal':
        candidates = [
            f"{id_clean}.jpg", f"{id_clean}.png", f"{id_clean}.JPG", f"{id_clean}.PNG", f"{id_clean}.jpeg",
            f"{id_padded}.jpg", f"{id_padded}.png", f"{id_padded}.JPG", f"{id_padded}.PNG", f"{id_padded}.jpeg"
        ]
    else:
        candidates = [
            f"{id_clean}_family.jpg", f"{id_clean}_family.png", f"{id_clean}_family.JPG", f"{id_clean}_family.PNG", f"{id_clean}_family.jpeg",
            f"{id_padded}_family.jpg", f"{id_padded}_family.png", f"{id_padded}_family.JPG", f"{id_padded}_family.PNG", f"{id_padded}_family.jpeg"
        ]
        
    for cand in candidates:
        full_path = os.path.join(PHOTOS_DIR, cand)
        if os.path.exists(full_path):
            return f"photos/{cand}"
    return ''

def import_excel():
    print(f"Loading Excel file: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    sheet = wb['Personnel Data']
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Re-create database table cleanly
    cursor.execute("DROP TABLE IF EXISTS military_personnel")
    cursor.execute('''
        CREATE TABLE military_personnel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            manual_id VARCHAR(50),
            rank VARCHAR(100),
            surname VARCHAR(100),
            given_name VARCHAR(100),
            name_khmer VARCHAR(150),
            name_latin VARCHAR(150),
            gender VARCHAR(10) DEFAULT 'ប',
            id_card VARCHAR(50),
            position VARCHAR(150),
            unit_group VARCHAR(150),
            unit VARCHAR(150),
            rank_date VARCHAR(50),
            position_date VARCHAR(50),
            dob VARCHAR(50),
            enlistment_date VARCHAR(50),
            framework_date VARCHAR(50),
            education_level VARCHAR(50),
            study_local VARCHAR(50),
            study_abroad VARCHAR(50),
            children_count VARCHAR(50),
            black_card_expiry VARCHAR(50),
            blue_card_expiry VARCHAR(50),
            pob_village VARCHAR(100),
            pob_commune VARCHAR(100),
            pob_district VARCHAR(100),
            pob_province VARCHAR(100),
            place_of_birth TEXT,
            addr_house VARCHAR(100),
            addr_group VARCHAR(100),
            addr_village VARCHAR(100),
            addr_commune VARCHAR(100),
            addr_district VARCHAR(100),
            addr_province VARCHAR(100),
            current_address TEXT,
            marital_status VARCHAR(50),
            phone VARCHAR(50),
            notes TEXT,
            photo TEXT,
            family_photo TEXT,
            family_name VARCHAR(100),
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    imported_count = 0
    photos_matched = 0
    family_photos_matched = 0
    
    current_section = ''

    for r in range(2, sheet.max_row + 1):
        vals = [sheet.cell(row=r, column=col).value for col in range(1, 28)]
        non_empty = [v for v in vals if v is not None and str(v).strip() != '']
        if len(non_empty) == 0:
            continue
        
        # Check section header
        if len(non_empty) <= 2 and vals[9] is not None:
            current_section = str(vals[9]).strip()
            continue

        manual_id = vals[1]
        rank = vals[2]
        surname = vals[3]
        given_name = vals[4]
        gender = vals[5]
        id_card = vals[6]
        position = vals[7]
        unit_group = vals[8]
        unit = vals[9]
        dob = vals[10]
        enlistment_date = vals[11]
        rank_date = vals[12]
        position_date = vals[13]
        education_level = vals[14]
        study_local_val = vals[15]
        study_abroad_val = vals[16]
        children_count = vals[17]
        phone = vals[18]
        photo_excel = vals[19]
        family_photo_excel = vals[20]
        family_name = vals[21]
        black_card_expiry = vals[22]
        blue_card_expiry = vals[23]
        place_of_birth = vals[24]
        current_address = vals[25]
        notes = vals[26]

        if not id_card and not surname and not given_name:
            continue

        id_card_str = str(id_card).strip() if id_card is not None else ''
        surname_str = str(surname).strip() if surname is not None else ''
        given_name_str = str(given_name).strip() if given_name is not None else ''
        name_khmer = f"{surname_str} {given_name_str}".strip()

        unit_str = str(unit).strip() if unit else current_section
        unit_group_str = str(unit_group).strip() if unit_group else 'លេខាធិការដ្ឋាន'

        # Photos
        photo_path = str(photo_excel).strip() if photo_excel else ''
        if photo_path and not os.path.exists(os.path.join(BASE_DIR, photo_path)):
            photo_path = ''
        if not photo_path:
            photo_path = find_photo_path(id_card_str, 'personal')
        if photo_path:
            photos_matched += 1

        family_photo_path = str(family_photo_excel).strip() if family_photo_excel else ''
        if family_photo_path and not os.path.exists(os.path.join(BASE_DIR, family_photo_path)):
            family_photo_path = ''
        if not family_photo_path:
            family_photo_path = find_photo_path(id_card_str, 'family')
        if family_photo_path:
            family_photos_matched += 1

        pob_v, pob_c, pob_d, pob_p = parse_pob(place_of_birth)
        addr_h, addr_g, addr_v, addr_c, addr_d, addr_p = parse_addr(current_address)

        cursor.execute('''
            INSERT INTO military_personnel (
                manual_id, rank, surname, given_name, name_khmer, gender, id_card, position,
                unit_group, unit, rank_date, position_date, dob, enlistment_date, framework_date,
                education_level, study_local, study_abroad, children_count, black_card_expiry, blue_card_expiry,
                pob_village, pob_commune, pob_district, pob_province, place_of_birth,
                addr_house, addr_group, addr_village, addr_commune, addr_district, addr_province, current_address,
                marital_status, phone, notes, photo, family_photo, family_name
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            str(manual_id).strip() if manual_id is not None else '',
            str(rank).strip() if rank is not None else '',
            surname_str,
            given_name_str,
            name_khmer,
            str(gender).strip() if gender is not None else 'ប',
            id_card_str,
            str(position).strip() if position is not None else '',
            unit_group_str,
            unit_str,
            format_date_str(rank_date),
            format_date_str(position_date),
            format_date_str(dob),
            format_date_str(enlistment_date),
            format_date_str(enlistment_date),
            str(education_level).strip() if education_level is not None else '',
            str(study_local_val).strip() if study_local_val is not None else '',
            str(study_abroad_val).strip() if study_abroad_val is not None else '',
            str(children_count).strip() if children_count is not None else '',
            format_date_str(black_card_expiry),
            format_date_str(blue_card_expiry),
            pob_v, pob_c, pob_d, pob_p,
            str(place_of_birth).strip() if place_of_birth is not None else '',
            addr_h, addr_g, addr_v, addr_c, addr_d, addr_p,
            str(current_address).strip() if current_address is not None else '',
            'រៀបការរួច' if family_name else 'នៅលីវ',
            str(phone).strip() if phone is not None else '',
            str(notes).strip() if notes is not None else '',
            photo_path,
            family_photo_path,
            str(family_name).strip() if family_name is not None else ''
        ))
        imported_count += 1

    conn.commit()
    conn.close()
    print(f" Successfully re-imported {imported_count} exact records from Personnel_Data.xlsx into SQLite database!")
    print(f" Personal Photos matched: {photos_matched}")
    print(f" Family Photos matched: {family_photos_matched}")

if __name__ == '__main__':
    import_excel()
